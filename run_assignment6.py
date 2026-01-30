import os
import sys
import time
import json
import textwrap
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions


import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas


# =========================
# CONFIG / CONSTANTS
# =========================
PROJECT_NAME = "Assignment 6 — SQAT"
TEST_SITE_NAME = "Booking.com"
DEFAULT_EXCEL = "testdata.xlsx"
DEFAULT_SHEET = "data"

# We'll run on BrowserStack (Windows 11) in 2 browsers: Chrome + Firefox
BROWSERS_TO_RUN = [
    {"browserName": "Chrome", "browserVersion": "latest"},
    {"browserName": "Firefox", "browserVersion": "latest"},
]

BSTACK_OS = {
    "os": "Windows",
    "osVersion": "11"
}


# =========================
# TASK 1: EXCEL TEST DATA
# =========================
def ensure_excel_exists(excel_path: Path, sheet_name: str) -> None:
    """
    Creates a sample Excel if not present, so your submission always has a reproducible dataset.
    """
    if excel_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row (key/value)
    ws["A1"] = "key"
    ws["B1"] = "value"

    # IMPORTANT: NO HARDCODED values in test logic; they come from Excel.
    # You may edit these values later.
    rows = [
        ("base_url", "https://www.booking.com/"),
        ("destination", "Astana"),
        ("checkin_offset_days", 14),   # check-in is today + offset
        ("checkout_offset_days", 17),  # check-out is today + offset
        ("adults", 2),
        ("rooms", 1),
        ("take_screenshots", "yes"),
        ("wait_timeout_sec", 25),
    ]

    for i, (k, v) in enumerate(rows, start=2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v

    wb.save(excel_path)


def read_test_data(excel_path: Path, sheet_name: str) -> dict:
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        value = row[1]
        data[key] = value

    # minimal validation
    required = ["base_url", "destination", "checkin_offset_days", "checkout_offset_days", "adults", "rooms"]
    missing = [k for k in required if k not in data]
    if missing:
        raise ValueError(f"Missing required keys in Excel: {missing}")

    return data


def make_data_proof_image(data: dict, out_png: Path) -> None:
    """
    Creates a PNG table image showing the Excel data that was read.
    This is convenient proof for the PDF report.
    """
    keys = list(data.keys())
    vals = [str(data[k]) for k in keys]

    fig = plt.figure(figsize=(9, max(3, 0.35 * len(keys) + 1)))
    ax = fig.add_subplot(111)
    ax.axis("off")

    table_data = [["key", "value"]] + [[k, v] for k, v in zip(keys, vals)]
    tbl = ax.table(cellText=table_data, loc="center")

    # Keep default matplotlib styling (no custom colors).
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(10)
    tbl.scale(1, 1.3)

    fig.tight_layout()
    fig.savefig(out_png, dpi=200)
    plt.close(fig)


# =========================
# TASK 2: BROWSERSTACK TEST
# =========================
def require_env(var_name: str) -> str:
    val = os.getenv(var_name)
    if not val:
        raise EnvironmentError(f"Missing env var: {var_name}")
    return val


def bstack_remote_url(username: str, access_key: str) -> str:
    return f"https://{username}:{access_key}@hub-cloud.browserstack.com/wd/hub"


def safe_screenshot(driver, path: Path) -> None:
    try:
        driver.save_screenshot(str(path))
    except Exception:
        # screenshot failure should not kill the whole run
        pass


def accept_cookies_if_present(driver, wait: WebDriverWait):
    """
    Booking.com sometimes shows cookie banner; handle best-effort.
    """
    candidates = [
        (By.ID, "onetrust-accept-btn-handler"),
        (By.CSS_SELECTOR, "button#onetrust-accept-btn-handler"),
        (By.XPATH, "//button[contains(.,'Accept')]"),
        (By.XPATH, "//button[contains(.,'I agree')]"),
    ]
    for by, sel in candidates:
        try:
            btn = wait.until(EC.element_to_be_clickable((by, sel)))
            btn.click()
            return True
        except Exception:
            continue
    return False


def booking_search_test(driver, data: dict, artifacts_dir: Path) -> dict:
    """
    Real website scenario: Booking.com hotel search using destination + dates.
    Assertions:
      - results page contains property cards or result container
    """
    timeout = int(data.get("wait_timeout_sec", 25))
    wait = WebDriverWait(driver, timeout)

    base_url = str(data["base_url"])
    destination = str(data["destination"])

    checkin_offset = int(data["checkin_offset_days"])
    checkout_offset = int(data["checkout_offset_days"])

    take_screens = str(data.get("take_screenshots", "yes")).lower() in ("yes", "true", "1")

    run_log = {
        "base_url": base_url,
        "destination": destination,
        "checkin_offset_days": checkin_offset,
        "checkout_offset_days": checkout_offset,
        "steps": [],
        "passed": False,
        "error": None,
    }

    def step(name: str):
        run_log["steps"].append({"time": datetime.now().isoformat(timespec="seconds"), "name": name})

    # Step 1: Open site
    step("Open Booking.com")
    driver.get(base_url)
    time.sleep(2)
    accept_cookies_if_present(driver, wait)
    if take_screens:
        safe_screenshot(driver, artifacts_dir / "01_home.png")

    # Step 2: Fill destination
    step("Fill destination")
    # Booking.com uses name="ss" often
    dest_input = wait.until(EC.element_to_be_clickable((By.NAME, "ss")))
    dest_input.clear()
    dest_input.send_keys(destination)
    time.sleep(1)
    dest_input.send_keys(Keys.ENTER)
    if take_screens:
        safe_screenshot(driver, artifacts_dir / "02_destination_filled.png")

    # Step 3: Pick dates
    step("Pick dates")
    # Booking uses date cell data-date="YYYY-MM-DD"
    today = datetime.now().date()
    checkin = (today.replace(day=today.day) + (datetime.now().date() - today))  # no-op, just clarity
    checkin = today.fromordinal(today.toordinal() + checkin_offset)
    checkout = today.fromordinal(today.toordinal() + checkout_offset)

    checkin_str = checkin.isoformat()
    checkout_str = checkout.isoformat()

    # Open calendar (best-effort)
    try:
        date_trigger = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "[data-testid='date-display-field-start']")))
        date_trigger.click()
    except Exception:
        # fallback: click anywhere on date widget
        try:
            driver.find_element(By.CSS_SELECTOR, "button[data-testid='searchbox-dates-container']").click()
        except Exception:
            pass

    # Click check-in and check-out cells
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"[data-date='{checkin_str}']"))).click()
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, f"[data-date='{checkout_str}']"))).click()

    if take_screens:
        safe_screenshot(driver, artifacts_dir / "03_dates_selected.png")

    # Step 4: Configure guests (adults/rooms) best-effort
    step("Configure guests")
    adults = int(data["adults"])
    rooms = int(data["rooms"])

    try:
        guest_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button[data-testid='occupancy-config']")))
        guest_btn.click()

        # Adults controls
        def set_counter(section_testid: str, desired: int):
            value_el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, f"div[data-testid='{section_testid}'] span[data-testid='counter-value']")))
            minus = driver.find_element(By.CSS_SELECTOR, f"div[data-testid='{section_testid}'] button[aria-label*='Decrease']")
            plus = driver.find_element(By.CSS_SELECTOR, f"div[data-testid='{section_testid}'] button[aria-label*='Increase']")

            def current():
                return int(value_el.text.strip())

            # Normalize to desired
            while current() > desired:
                minus.click()
                time.sleep(0.3)
            while current() < desired:
                plus.click()
                time.sleep(0.3)

        set_counter("occupancy-adults", adults)
        set_counter("occupancy-rooms", rooms)

        if take_screens:
            safe_screenshot(driver, artifacts_dir / "04_guests_set.png")
    except Exception:
        # Not fatal; site UI can vary
        run_log["steps"].append({"time": datetime.now().isoformat(timespec="seconds"), "name": "Guests UI not adjusted (UI variation)"})


    # Step 5: Submit search
    step("Submit search")
    search_btn_selectors = [
        (By.CSS_SELECTOR, "button[type='submit']"),
        (By.CSS_SELECTOR, "button[data-testid='searchbox-submit-button']"),
    ]
    clicked = False
    for by, sel in search_btn_selectors:
        try:
            btn = wait.until(EC.element_to_be_clickable((by, sel)))
            btn.click()
            clicked = True
            break
        except Exception:
            continue
    if not clicked:
        raise RuntimeError("Could not find/click search button")

    # Step 6: Assert results
    step("Wait for results and assert")
    # Common result container/card testids:
    result_candidates = [
        (By.CSS_SELECTOR, "[data-testid='property-card']"),
        (By.CSS_SELECTOR, "div[data-testid='property-card-container']"),
        (By.CSS_SELECTOR, "#search_results_table"),
    ]

    found = False
    for by, sel in result_candidates:
        try:
            wait.until(EC.presence_of_element_located((by, sel)))
            found = True
            break
        except Exception:
            continue

    if take_screens:
        safe_screenshot(driver, artifacts_dir / "05_results.png")

    if not found:
        raise AssertionError("Results were not detected (no property cards/container found).")

    run_log["passed"] = True
    run_log["checkin"] = checkin_str
    run_log["checkout"] = checkout_str
    return run_log


def run_on_browserstack(data: dict, artifacts_root: Path) -> list[dict]:
    username = require_env("BROWSERSTACK_USERNAME")
    access_key = require_env("BROWSERSTACK_ACCESS_KEY")
    remote_url = bstack_remote_url(username, access_key)

    results = []

    for b in BROWSERS_TO_RUN:
        browser_name = b["browserName"]
        browser_version = b["browserVersion"]

        run_dir = artifacts_root / f"{BSTACK_OS['os']}_{BSTACK_OS['osVersion']}_{browser_name}"
        run_dir.mkdir(parents=True, exist_ok=True)

        caps = {
            "bstack:options": {
                "os": BSTACK_OS["os"],
                "osVersion": BSTACK_OS["osVersion"],
                "sessionName": f"Assignment6-{browser_name}",
                "buildName": f"Assignment6-{datetime.now().strftime('%Y%m%d')}",
                # You can add: "projectName": "SQAT"
            },
            "browserName": browser_name,
            "browserVersion": browser_version,
        }

        driver = None
        run_log = {
            "browser": browser_name,
            "browserVersion": browser_version,
            "os": BSTACK_OS["os"],
            "osVersion": BSTACK_OS["osVersion"],
            "passed": False,
            "error": None,
            "artifacts_dir": str(run_dir),
        }

        try:
            # Selenium 4 + W3C: use Options and set_capability
            if browser_name.lower() == "chrome":
                options = ChromeOptions()
            elif browser_name.lower() == "firefox":
                options = FirefoxOptions()
            else:
                raise ValueError(f"Unsupported browser: {browser_name}")

            # Base capabilities
            options.set_capability("browserName", browser_name)
            options.set_capability("browserVersion", browser_version)

            # BrowserStack-specific capabilities
            options.set_capability("bstack:options", {
                "os": BSTACK_OS["os"],
                "osVersion": BSTACK_OS["osVersion"],
                "sessionName": f"Assignment6-{browser_name}",
                "buildName": f"Assignment6-{datetime.now().strftime('%Y%m%d')}",
            })

            driver = webdriver.Remote(command_executor=remote_url, options=options)
            driver.set_window_size(1366, 768)


            test_log = booking_search_test(driver, data, run_dir)
            run_log.update(test_log)
            run_log["passed"] = True

        except Exception as e:
            run_log["passed"] = False
            run_log["error"] = f"{type(e).__name__}: {e}"
            # save final screenshot if possible
            if driver is not None:
                safe_screenshot(driver, run_dir / "99_error.png")
            run_log["traceback"] = traceback.format_exc()

        finally:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    pass

        # also store JSON
        with open(run_dir / "run_log.json", "w", encoding="utf-8") as f:
            json.dump(run_log, f, ensure_ascii=False, indent=2)

        results.append(run_log)

    return results


# =========================
# PDF REPORT GENERATION
# =========================
def read_own_code_snippet(py_file: Path, max_lines: int = 220) -> str:
    """
    Returns a snippet of this script (as evidence/code in PDF).
    """
    lines = py_file.read_text(encoding="utf-8").splitlines()
    snippet = lines[:max_lines]
    return "\n".join(snippet)


def draw_wrapped_text(c: canvas.Canvas, text: str, x: float, y: float, width: float, line_height: float):
    """
    Draw multi-line wrapped text; returns new y.
    """
    for paragraph in text.split("\n"):
        if paragraph.strip() == "":
            y -= line_height
            continue
        wrapped = textwrap.wrap(paragraph, width=int(width // 6.2))  # heuristic
        for line in wrapped:
            c.drawString(x, y, line)
            y -= line_height
    return y


def add_image_if_exists(c: canvas.Canvas, img_path: Path, x: float, y: float, max_w: float, max_h: float, caption: str | None = None):
    if not img_path.exists():
        return y
    try:
        img = ImageReader(str(img_path))
        iw, ih = img.getSize()
        scale = min(max_w / iw, max_h / ih)
        w = iw * scale
        h = ih * scale
        c.drawImage(img, x, y - h, width=w, height=h, preserveAspectRatio=True, mask='auto')
        y = y - h - 6
        if caption:
            c.setFont("Helvetica", 9)
            c.drawString(x, y, caption)
            c.setFont("Helvetica", 10)
            y -= 14
        else:
            y -= 10
    except Exception:
        pass
    return y


def generate_pdf_report(
    out_pdf: Path,
    artifacts_root: Path,
    data_png: Path,
    excel_path: Path,
    run_results: list[dict],
    code_snippet: str,
):
    c = canvas.Canvas(str(out_pdf), pagesize=A4)
    w, h = A4
    margin = 18 * mm
    x = margin
    y = h - margin

    # Title
    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, PROJECT_NAME)
    y -= 22
    c.setFont("Helvetica", 11)
    c.drawString(x, y, f"Student environment: macOS (local runner), Remote execution: BrowserStack")
    y -= 16
    c.drawString(x, y, f"Website under test: {TEST_SITE_NAME}")
    y -= 16
    c.drawString(x, y, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 24

    # Task overview
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Overview")
    y -= 16
    c.setFont("Helvetica", 10)
    overview = (
        "Task 1: Read test data from Excel using openpyxl (no hardcoded test values).\n"
        "Task 2: Run Selenium tests on BrowserStack using Remote WebDriver on at least 2 browsers and 1 OS.\n"
        "Evidence included: code snippet, proof of Excel data read, screenshots of execution, and short explanations."
    )
    y = draw_wrapped_text(c, overview, x, y, width=w - 2 * margin, line_height=13)
    y -= 8

    # Task 1 section
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Task 1 — Test data from Excel (openpyxl)")
    y -= 16
    c.setFont("Helvetica", 10)
    task1_text = (
        f"Excel file used: {excel_path.name}\n"
        f"Sheet: {DEFAULT_SHEET}\n"
        "All input values (URL, destination, date offsets, guests) are read from Excel.\n"
        "The automation logic does not contain hardcoded test values."
    )
    y = draw_wrapped_text(c, task1_text, x, y, width=w - 2 * margin, line_height=13)
    y -= 6

    y = add_image_if_exists(
        c,
        data_png,
        x=x,
        y=y,
        max_w=w - 2 * margin,
        max_h=120 * mm,
        caption="Figure 1 — Proof image: data read from Excel and used by the test",
    )

    # New page if needed
    if y < 120 * mm:
        c.showPage()
        y = h - margin
        c.setFont("Helvetica", 10)

    # Task 2 section
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Task 2 — BrowserStack execution (Selenium Remote WebDriver)")
    y -= 16
    c.setFont("Helvetica", 10)
    browsers_list = ", ".join([f"{r.get('browser')} ({r.get('browserVersion')})" for r in run_results])
    task2_text = (
        f"Chosen OS: {BSTACK_OS['os']} {BSTACK_OS['osVersion']}\n"
        f"Browsers: {browsers_list}\n"
        "Rationale: running on two different browsers helps detect browser-specific UI/DOM differences.\n"
        "Screenshots are captured at key steps: home page, destination filled, dates selected, guests set (if available), results page."
    )
    y = draw_wrapped_text(c, task2_text, x, y, width=w - 2 * margin, line_height=13)
    y -= 8

    # Add a couple screenshots from each run (if exist)
    for r in run_results:
        run_dir = Path(r["artifacts_dir"])
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x, y, f"Run: {r.get('os')} {r.get('osVersion')} — {r.get('browser')} ({r.get('browserVersion')}) — PASSED={r.get('passed')}")
        y -= 14
        c.setFont("Helvetica", 10)

        shots = [
            ("01_home.png", "Home page opened"),
            ("03_dates_selected.png", "Dates selected"),
            ("05_results.png", "Results page"),
            ("99_error.png", "Error screenshot (if failed)"),
        ]
        for fn, cap in shots:
            y = add_image_if_exists(
                c,
                run_dir / fn,
                x=x,
                y=y,
                max_w=w - 2 * margin,
                max_h=70 * mm,
                caption=f"{run_dir.name}: {cap}",
            )
            if y < 90 * mm:
                c.showPage()
                y = h - margin
                c.setFont("Helvetica", 10)

        # brief status text
        status = "PASSED" if r.get("passed") else f"FAILED: {r.get('error')}"
        y = draw_wrapped_text(c, f"Status: {status}", x, y, width=w - 2 * margin, line_height=13)
        y -= 10
        if y < 80 * mm:
            c.showPage()
            y = h - margin
            c.setFont("Helvetica", 10)

    # Code snippet section
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "Code snippet (evidence)")
    y -= 16
    c.setFont("Helvetica", 8)

    # Draw code in chunks per page
    code_lines = code_snippet.splitlines()
    lines_per_page = 65
    i = 0
    while i < len(code_lines):
        if y < 30 * mm:
            c.showPage()
            y = h - margin
            c.setFont("Helvetica", 8)

        chunk = code_lines[i:i + lines_per_page]
        for line in chunk:
            c.drawString(x, y, line[:130])
            y -= 10
            if y < 25 * mm:
                break
        i += lines_per_page

    c.save()


# =========================
# MAIN
# =========================
def main():
    here = Path(__file__).resolve().parent
    excel_path = here / DEFAULT_EXCEL
    ensure_excel_exists(excel_path, DEFAULT_SHEET)

    artifacts_root = here / "artifacts" / datetime.now().strftime("%Y%m%d_%H%M%S")
    artifacts_root.mkdir(parents=True, exist_ok=True)

    # Task 1: Read Excel
    data = read_test_data(excel_path, DEFAULT_SHEET)

    # Create proof image for PDF (shows Excel data read)
    data_png = artifacts_root / "excel_data_proof.png"
    make_data_proof_image(data, data_png)

    # Task 2: Run BrowserStack tests
    run_results = run_on_browserstack(data, artifacts_root)

    # Generate PDF report
    out_pdf = artifacts_root / "Assignment6_Report.pdf"
    code_snippet = read_own_code_snippet(Path(__file__))
    generate_pdf_report(
        out_pdf=out_pdf,
        artifacts_root=artifacts_root,
        data_png=data_png,
        excel_path=excel_path,
        run_results=run_results,
        code_snippet=code_snippet,
    )

    print("\n=== DONE ===")
    print(f"Artifacts folder: {artifacts_root}")
    print(f"PDF report: {out_pdf}")
    print("Screenshots are inside each browser run folder.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("FATAL ERROR:", e)
        sys.exit(1)
